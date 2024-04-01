import pandas as pd
import openpyxl
import os

from openpyxl.styles import Font, Alignment, Side, Border, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

import pandas as pd


# Formating Function
def apply_formatting(worksheet, df):
    # Apply formatting to the entire table
    for row in worksheet.iter_rows(min_row=1, max_row=len(df) + 1, max_col=len(df.columns)):
        for cell in row:
            cell.font = Font(name='Times New Roman', size=10)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))
    
    # Apply formatting to the header row
    header_row = worksheet[1]
    for cell in header_row:
        cell.font = Font(name='Times New Roman', size=10, bold=True)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        
    # Resize column widths to fit the content of headers and set maximum width to 100
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 100)
        worksheet.column_dimensions[column].width = adjusted_width

    # Freeze the header row
    worksheet.freeze_panes = worksheet['A2']
    
    # Add filters to the header row
    worksheet.auto_filter.ref = worksheet.dimensions




# Function to shade empty cells with a specified fill color
def shade_white_cells(worksheet, df, fill_color='F2F2F2'):
    fill_color = 'FF' + fill_color  # Adding alpha value to the color
    for index, row in df.iterrows():
        for col_index, value in enumerate(row):
            if pd.isnull(value) or value == '':
                cell = worksheet.cell(row=index + 2, column=col_index + 1)  # Adding 2 to index for header row offset
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

'''
import pandas as pd
import openpyxl
import os

def write_dataframe_to_excel(df, file_path, sheet_name):
    # Check if the file exists
    file_exists = os.path.isfile(file_path)
    
    # Open the Excel writer in append mode if the file exists, otherwise create a new file
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a' if file_exists else 'w') as writer:
        # If the file doesn't exist, create a new workbook
        if not file_exists:
            writer.book = openpyxl.Workbook()
            # Make the sheet visible
            writer.book.active.sheet_state = 'visible'
        
        # Try to load existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        
        # Write DataFrame to Excel
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        
        # Apply formatting if necessary
        if sheet_name in writer.sheets:
            apply_formatting(writer.sheets[sheet_name], df)
        
        # Save changes
        writer.save()



'''
# Function to write the results to Excel
def write_dataframe_to_excel(df, file_path, sheet_name):
    # Check if the file exists
    file_exists = os.path.isfile(file_path)
    
    if file_exists:
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            apply_formatting(worksheet, df)
            # shade_white_cells(worksheet, df)
    else:
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            apply_formatting(worksheet, df)
            # shade_white_cells(worksheet, df)
            


# list of LOOs
table_data = [
    {"File Name": "Commercial-Line-Haul", "Process Number": "09-CLH", "Process Scenario": "09- Commercial Line Haul"},
    {"File Name": "Carrier_Function", "Process Number": "01-CF", "Process Scenario": "01- Carrier Function"},
    {"File Name": "Rate_Management", "Process Number": "02-RM", "Process Scenario": "02- Rate Management"},
    {"File Name": "Payment_Process", "Process Number": "03-TPPS", "Process Scenario": "03- Third Party Payment System"},
    {"File Name": "Deployment", "Process Number": "04-DEP", "Process Scenario": "04- Deployment"},
    {"File Name": "Rail", "Process Number": "05-RAIL", "Process Scenario": "05- Rail"},
    {"File Name": "Inbound_Terminating_Freight", "Process Number": "06-ITF", "Process Scenario": "06- Inbound Terminating Freight"},
    {"File Name": "Support_HQ", "Process Number": "07-SUP", "Process Scenario": "07- Support HQ: Carrier Performance"},
    {"File Name": "Outbound", "Process Number": "08-OBF", "Process Scenario": "08- Outbound Freight"},
    {"File Name": "HAZMAT", "Process Number": "10-HZMT", "Process Scenario": "10- Hazmat"},
    {"File Name": "TPS", "Process Number": "11-TPS", "Process Scenario": "11-TPS"},
    {"File Name": "ITV", "Process Number": "12-ITV", "Process Scenario": "12-ITV"},
    {"File Name": "TAC", "Process Number": "13-TAC", "Process Scenario": "13-TAC"}
]